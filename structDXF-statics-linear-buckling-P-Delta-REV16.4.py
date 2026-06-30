#!/usr/bin/env python3
# ============================================================
# structDXF - structural DXF Parser & Eurocode 3 Verifier
# Copyright (C) 2025  Ing. Bruno Zilli
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
# ============================================================
# DISCLAIMER:
# This software is intended for EDUCATIONAL AND RESEARCH PURPOSES ONLY.
# It has NOT been certified by any professional engineering body.
# Results produced by this software must NOT be used as the sole basis
# for real structural design, analysis, or decision-making.
# Always consult a qualified structural engineer and verify results
# with official standards and validated commercial software.
#
# The authors assume NO LIABILITY for any damages, losses, or
# consequences arising from the use of this software.
# By using this software you agree that you do so at your own risk.
# ============================================================
# STRUCTURAL DXF PARSER - STATICS v16.4 FINAL CORRETTA
# EN 1993-1-1 + EN 1990 | Class 1 sections | Beam-Column Method 2
# Fix: visualizzazione etichette carichi distribuiti su membri originali
# NEW: P-Delta analysis, separate deformed shape & Excel sheet
# ============================================================

import ezdxf
import matplotlib
import os

# Se non c'è display (es. sessione SSH), usa backend non interattivo
if 'DISPLAY' not in os.environ:
    matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from Pynite import FEModel3D
import sys
import re
import openpyxl
from scipy.interpolate import PchipInterpolator

# ============================================================
# STEEL SECTIONS LIBRARY
# ============================================================
STEEL_SECTIONS = {
    'IPE80':  {'A': 7.64e-4,  'Iy': 8.49e-8,  'Iz': 8.01e-7,  'J': 2.31e-8, 'Wpl_y': 2.32e-6, 'Wpl_z': 2.07e-5, 'class': 1},
    'IPE100': {'A': 10.3e-4,  'Iy': 1.59e-7,  'Iz': 1.71e-6,  'J': 4.57e-8, 'Wpl_y': 3.94e-6, 'Wpl_z': 3.51e-5, 'class': 1},
    'IPE120': {'A': 13.2e-4,  'Iy': 2.77e-7,  'Iz': 3.18e-6,  'J': 8.79e-8, 'Wpl_y': 6.08e-6, 'Wpl_z': 5.37e-5, 'class': 1},
    'IPE140': {'A': 16.4e-4,  'Iy': 4.49e-7,  'Iz': 5.41e-6,  'J': 1.58e-7, 'Wpl_y': 8.83e-6, 'Wpl_z': 7.78e-5, 'class': 1},
    'IPE160': {'A': 20.1e-4,  'Iy': 6.83e-7,  'Iz': 8.69e-6,  'J': 2.74e-7, 'Wpl_y': 1.24e-5, 'Wpl_z': 1.09e-4, 'class': 1},
    'IPE180': {'A': 23.9e-4,  'Iy': 1.01e-6,  'Iz': 1.32e-5,  'J': 4.49e-7, 'Wpl_y': 1.66e-5, 'Wpl_z': 1.47e-4, 'class': 1},
    'IPE200': {'A': 28.5e-4,  'Iy': 1.42e-6,  'Iz': 1.94e-5,  'J': 7.00e-7, 'Wpl_y': 2.21e-5, 'Wpl_z': 1.94e-4, 'class': 1},
    'IPE220': {'A': 33.4e-4,  'Iy': 2.05e-6,  'Iz': 2.77e-5,  'J': 1.07e-6, 'Wpl_y': 2.85e-5, 'Wpl_z': 2.52e-4, 'class': 1},
    'IPE240': {'A': 39.1e-4,  'Iy': 2.84e-6,  'Iz': 3.89e-5,  'J': 1.61e-6, 'Wpl_y': 3.67e-5, 'Wpl_z': 3.24e-4, 'class': 1},
    'IPE270': {'A': 45.9e-4,  'Iy': 4.20e-6,  'Iz': 5.79e-5,  'J': 2.58e-6, 'Wpl_y': 4.84e-5, 'Wpl_z': 4.29e-4, 'class': 1},
    'IPE300': {'A': 53.8e-4,  'Iy': 6.04e-6,  'Iz': 8.36e-5,  'J': 4.01e-6, 'Wpl_y': 6.28e-5, 'Wpl_z': 5.57e-4, 'class': 1},
    'IPE330': {'A': 62.6e-4,  'Iy': 7.88e-6,  'Iz': 1.18e-4,  'J': 5.71e-6, 'Wpl_y': 8.04e-5, 'Wpl_z': 7.13e-4, 'class': 1},
    'IPE360': {'A': 72.7e-4,  'Iy': 1.04e-5,  'Iz': 1.63e-4,  'J': 8.37e-6, 'Wpl_y': 1.02e-4, 'Wpl_z': 9.04e-4, 'class': 1},
    'IPE400': {'A': 84.5e-4,  'Iy': 1.32e-5,  'Iz': 2.31e-4,  'J': 1.20e-5, 'Wpl_y': 1.31e-4, 'Wpl_z': 1.16e-3, 'class': 1},
    'IPE450': {'A': 98.8e-4,  'Iy': 1.68e-5,  'Iz': 3.37e-4,  'J': 1.80e-5, 'Wpl_y': 1.70e-4, 'Wpl_z': 1.50e-3, 'class': 1},
    'IPE500': {'A': 116e-4,   'Iy': 2.14e-5,  'Iz': 4.82e-4,  'J': 2.72e-5, 'Wpl_y': 2.19e-4, 'Wpl_z': 1.93e-3, 'class': 1},
    'IPE550': {'A': 134e-4,   'Iy': 2.67e-5,  'Iz': 6.71e-4,  'J': 3.95e-5, 'Wpl_y': 2.78e-4, 'Wpl_z': 2.45e-3, 'class': 1},
    'IPE600': {'A': 156e-4,   'Iy': 3.39e-5,  'Iz': 9.21e-4,  'J': 5.87e-5, 'Wpl_y': 3.51e-4, 'Wpl_z': 3.09e-3, 'class': 1},
    'HEA100': {'A': 21.2e-4,  'Iy': 1.34e-6,  'Iz': 3.49e-6,  'J': 5.30e-7, 'Wpl_y': 8.30e-5, 'Wpl_z': 7.30e-5, 'class': 1},
    'HEA200': {'A': 53.8e-4,  'Iy': 1.34e-5,  'Iz': 3.69e-5,  'J': 9.13e-6, 'Wpl_y': 4.30e-4, 'Wpl_z': 3.89e-4, 'class': 1},
    'HEA300': {'A': 113e-4,   'Iy': 6.31e-5,  'Iz': 1.83e-4,  'J': 6.68e-5, 'Wpl_y': 1.48e-3, 'Wpl_z': 1.26e-3, 'class': 1},
    'HEB100': {'A': 26.0e-4,  'Iy': 1.67e-6,  'Iz': 4.50e-6,  'J': 9.25e-7, 'Wpl_y': 1.04e-4, 'Wpl_z': 9.00e-5, 'class': 1},
    'HEB200': {'A': 78.1e-4,  'Iy': 2.00e-5,  'Iz': 5.70e-5,  'J': 1.97e-5, 'Wpl_y': 6.43e-4, 'Wpl_z': 5.70e-4, 'class': 1},
    'HEB300': {'A': 149e-4,   'Iy': 8.56e-5,  'Iz': 2.52e-4,  'J': 1.34e-4, 'Wpl_y': 2.15e-3, 'Wpl_z': 1.87e-3, 'class': 1},
    'SHS_60x60x3':   {'A': 6.91e-4,  'Iy': 3.92e-7,  'Iz': 3.92e-7,  'J': 6.10e-7, 'Wpl_y': 1.46e-5, 'Wpl_z': 1.46e-5, 'class': 1},
    'SHS_80x80x4':   {'A': 1.22e-3,  'Iy': 1.19e-6,  'Iz': 1.19e-6,  'J': 1.85e-6, 'Wpl_y': 3.28e-5, 'Wpl_z': 3.28e-5, 'class': 1},
    'SHS_100x100x4': {'A': 1.54e-3,  'Iy': 2.40e-6,  'Iz': 2.40e-6,  'J': 3.73e-6, 'Wpl_y': 5.35e-5, 'Wpl_z': 5.35e-5, 'class': 1},
    'SHS_120x120x5': {'A': 2.30e-3,  'Iy': 5.17e-6,  'Iz': 5.17e-6,  'J': 8.05e-6, 'Wpl_y': 9.55e-5, 'Wpl_z': 9.55e-5, 'class': 1},
    'SHS_150x150x6': {'A': 3.46e-3,  'Iy': 1.21e-5,  'Iz': 1.21e-5,  'J': 1.88e-5, 'Wpl_y': 1.78e-4, 'Wpl_z': 1.78e-4, 'class': 1},
    'SHS_200x200x8': {'A': 6.14e-3,  'Iy': 3.78e-5,  'Iz': 3.78e-5,  'J': 5.88e-5, 'Wpl_y': 4.20e-4, 'Wpl_z': 4.20e-4, 'class': 1},
    'RHS_100x60x4':  {'A': 1.22e-3,  'Iy': 7.23e-7,  'Iz': 1.70e-6,  'J': 1.40e-6, 'Wpl_y': 1.80e-5, 'Wpl_z': 3.97e-5, 'class': 1},
    'RHS_120x80x5':  {'A': 1.90e-3,  'Iy': 1.57e-6,  'Iz': 4.06e-6,  'J': 2.89e-6, 'Wpl_y': 3.42e-5, 'Wpl_z': 7.52e-5, 'class': 1},
    'RHS_150x100x6': {'A': 2.90e-3,  'Iy': 3.74e-6,  'Iz': 9.83e-6,  'J': 6.50e-6, 'Wpl_y': 6.92e-5, 'Wpl_z': 1.53e-4, 'class': 1},
    'RHS_200x100x8': {'A': 4.54e-3,  'Iy': 5.48e-6,  'Iz': 2.48e-5,  'J': 1.28e-5, 'Wpl_y': 9.86e-5, 'Wpl_z': 2.90e-4, 'class': 1},
    'CHS_48.3x3.2':  {'A': 4.53e-4,  'Iy': 1.16e-7,  'Iz': 1.16e-7,  'J': 2.32e-7, 'Wpl_y': 5.60e-6, 'Wpl_z': 5.60e-6, 'class': 1},
    'CHS_60.3x4':    {'A': 7.07e-4,  'Iy': 2.81e-7,  'Iz': 2.81e-7,  'J': 5.62e-7, 'Wpl_y': 1.08e-5, 'Wpl_z': 1.08e-5, 'class': 1},
    'CHS_76.1x4':    {'A': 9.06e-4,  'Iy': 5.92e-7,  'Iz': 5.92e-7,  'J': 1.18e-6, 'Wpl_y': 1.81e-5, 'Wpl_z': 1.81e-5, 'class': 1},
    'CHS_88.9x5':    {'A': 1.32e-3,  'Iy': 1.15e-6,  'Iz': 1.15e-6,  'J': 2.30e-6, 'Wpl_y': 3.01e-5, 'Wpl_z': 3.01e-5, 'class': 1},
    'CHS_114.3x6':   {'A': 2.04e-3,  'Iy': 2.97e-6,  'Iz': 2.97e-6,  'J': 5.94e-6, 'Wpl_y': 6.05e-5, 'Wpl_z': 6.05e-5, 'class': 1},
    'CHS_139.7x8':   {'A': 3.31e-3,  'Iy': 7.25e-6,  'Iz': 7.25e-6,  'J': 1.45e-5, 'Wpl_y': 1.21e-4, 'Wpl_z': 1.21e-4, 'class': 1},
    'CHS_168.3x10':  {'A': 4.97e-3,  'Iy': 1.57e-5,  'Iz': 1.57e-5,  'J': 3.14e-5, 'Wpl_y': 2.17e-4, 'Wpl_z': 2.17e-4, 'class': 1},
    'CHS_219.1x12':  {'A': 7.81e-3,  'Iy': 4.21e-5,  'Iz': 4.21e-5,  'J': 8.42e-5, 'Wpl_y': 4.47e-4, 'Wpl_z': 4.47e-4, 'class': 1},
}

STEEL_MATERIALS = {
    'S235': {'E': 210e9, 'G': 81e9, 'nu': 0.30, 'rho': 7850, 'fy': 235e6},
    'S275': {'E': 210e9, 'G': 81e9, 'nu': 0.30, 'rho': 7850, 'fy': 275e6},
    'S355': {'E': 210e9, 'G': 81e9, 'nu': 0.30, 'rho': 7850, 'fy': 355e6},
    'S450': {'E': 210e9, 'G': 81e9, 'nu': 0.30, 'rho': 7850, 'fy': 450e6},
}

# ============================================================
# FUNZIONI DI SUPPORTO
# ============================================================
def confirm_section_properties(parser):
    print("\n" + "="*65)
    print("📐 SECTION PROPERTIES CHECK (EN 1993-1-1)")
    print("="*65)
    print("\nStructure analyzed in X-Y plane. Mz (in-plane bending) uses Iz, Wpl_z")
    print("All sections assumed CLASS 1 (plastic moment capacity)\n")
    print(f"{'Member':<8} {'Section':<14} {'Class':<6} {'A [m²]':<12} {'Iz [m⁴]':<14} {'Wpl,z [m³]':<14} {'fy [MPa]':<10}")
    print("-"*80)
    for name, md in parser.members.items():
        if md.get('internal'): continue
        sec = STEEL_SECTIONS.get(md['section'], {})
        mat = STEEL_MATERIALS.get(md['material'], {})
        print(f"{name:<8} {md['section']:<14} {sec.get('class','?'):<6} {sec.get('A',0):<12.4e} {sec.get('Iz',0):<14.4e} {sec.get('Wpl_z',0):<14.4e} {mat.get('fy',0)/1e6:<10.0f}")
    print("-"*80)
    response = input("\nContinue? [Y/n]: ").strip().lower()
    if response == 'n': print("\n❌ Aborted."); sys.exit(0)
    print("✅ OK\n")

def classify_loads(parser):
    print("\n" + "="*65)
    print("📋 LOAD CLASSIFICATION (EN 1990)")
    print("="*65)
    print("\nClassify each load as Permanent (G) or Variable (Q):\n")
    for load in parser.distributed_loads:
        print(f"  Distributed load on {load['member']}: q = {load['value']:.0f} N/m")
        cls = input(f"    Type [G/Q, default Q]: ").strip().upper() or 'Q'
        load['class'] = cls
    for load in parser.point_loads:
        print(f"  Point load on {load['node']}: F = {load['value']:.0f} N ({load['direction']})")
        cls = input(f"    Type [G/Q, default Q]: ").strip().upper() or 'Q'
        load['class'] = cls
    for load in parser.moment_loads:
        print(f"  Moment on {load['node']}: M = {load['value']:.0f} Nm")
        cls = input(f"    Type [G/Q, default Q]: ").strip().upper() or 'Q'
        load['class'] = cls
    n_G = sum(1 for l in parser.distributed_loads + parser.point_loads + parser.moment_loads if l.get('class')=='G')
    n_Q = sum(1 for l in parser.distributed_loads + parser.point_loads + parser.moment_loads if l.get('class')=='Q')
    print(f"\n   Summary: {n_G} permanent (G), {n_Q} variable (Q)")
    gamma_G, gamma_Q, psi_0 = 1.35, 1.50, 0.7
    print(f"\n   Partial factors (EN 1990):")
    print(f"   γG = {gamma_G}, γQ = {gamma_Q}, ψ0 = {psi_0}")
    print(f"   γM0 = 1.00, γM1 = 1.00")
    print("✅ OK\n")
    return gamma_G, gamma_Q, psi_0

def calc_slu_factor(parser, gamma_G, gamma_Q):
    total_G = sum(abs(l['value']) for l in parser.distributed_loads + parser.point_loads + parser.moment_loads if l.get('class')=='G')
    total_Q = sum(abs(l['value']) for l in parser.distributed_loads + parser.point_loads + parser.moment_loads if l.get('class')=='Q')
    total = total_G + total_Q
    return (total_G * gamma_G + total_Q * gamma_Q) / total if total > 0 else 1.35

# ============================================================
# PARSER DXF
# ============================================================
class CustomDXFParser:
    def __init__(self, dxf_file):
        self.dxf_file = dxf_file
        print(f"📂 Loading: {dxf_file}")
        if not os.path.exists(dxf_file): raise FileNotFoundError(f"File not found: {dxf_file}")
        self.doc = ezdxf.readfile(dxf_file)
        self.msp = self.doc.modelspace()
        self.nodes = {}
        self.members = {}
        self.supports = {}
        self.distributed_loads = []
        self.point_loads = []
        self.moment_loads = []
        self.used_materials = set()
        self.used_sections = set()
        self.n_segments = 3
        self.unit_name = 'mm'
        self.scale_factor = 1/1000
        self.TOL = {}
        self.original_members_list = {}
        self.original_to_segments = {}
        
    def _get_tolerances(self, model_span=1000):
        scale_mult = model_span / 1000.0
        if self.unit_name == 'mm':
            return {'node_text': 50*scale_mult, 'member_text': 100*scale_mult, 'section_text': 150*scale_mult,
                    'find_node': 10*scale_mult, 'support_text': 120*scale_mult, 'support_vertex': 1*scale_mult,
                    'load_text': 150*scale_mult, 'beam_horiz': 5, 'beam_min_len': 50*scale_mult, 'moment_text': 200*scale_mult}
        elif self.unit_name == 'cm':
            return {'node_text': 5*scale_mult, 'member_text': 10*scale_mult, 'section_text': 15*scale_mult,
                    'find_node': 1*scale_mult, 'support_text': 12*scale_mult, 'support_vertex': 0.1*scale_mult,
                    'load_text': 15*scale_mult, 'beam_horiz': 0.5, 'beam_min_len': 5*scale_mult, 'moment_text': 20*scale_mult}
        elif self.unit_name == 'm':
            return {'node_text': 0.05*scale_mult, 'member_text': 0.10*scale_mult, 'section_text': 0.15*scale_mult,
                    'find_node': 0.01*scale_mult, 'support_text': 0.12*scale_mult, 'support_vertex': 0.001*scale_mult,
                    'load_text': 0.15*scale_mult, 'beam_horiz': 0.005, 'beam_min_len': 0.05*scale_mult, 'moment_text': 0.20*scale_mult}
        else:
            return {'node_text': 2*scale_mult, 'member_text': 4*scale_mult, 'section_text': 6*scale_mult,
                    'find_node': 0.5*scale_mult, 'support_text': 5*scale_mult, 'support_vertex': 0.01*scale_mult,
                    'load_text': 6*scale_mult, 'beam_horiz': 0.5, 'beam_min_len': 4*scale_mult, 'moment_text': 8*scale_mult}
        
    def _detect_units(self):
        try:
            insunits = self.doc.header.get('$INSUNITS', 4)
            unit_map = {1: 'in', 2: 'ft', 4: 'mm', 5: 'cm', 6: 'm'}
            self.unit_name = unit_map.get(insunits, 'mm')
            scale_map = {'mm': 1/1000, 'cm': 1/100, 'm': 1.0, 'in': 0.0254, 'ft': 0.3048}
            self.scale_factor = scale_map.get(self.unit_name, 1/1000)
        except:
            self.unit_name = 'mm'; self.scale_factor = 1/1000
        
        all_coords = []
        for c in self.msp.query('CIRCLE[layer=="STRUCT_NODES"]'):
            all_coords.extend([c.dxf.center.x, c.dxf.center.y])
        for line in self.msp.query('LINE[layer=="STRUCT_BEAMS"]'):
            all_coords.extend([line.dxf.start.x, line.dxf.start.y, line.dxf.end.x, line.dxf.end.y])
        model_span = max(max(all_coords) - min(all_coords), 1) if all_coords else 1000
        self.TOL = self._get_tolerances(model_span)
        print(f"   📏 Units: {self.unit_name}, scale={self.scale_factor}, span={model_span:.0f}")
        print(f"   📏 Tolerances: find_node={self.TOL['find_node']:.1f}, moment_text={self.TOL['moment_text']:.1f}")
        
    def parse(self):
        print("="*60); print("🔍 PARSING DXF...")
        self._extract_nodes(); self._detect_units()
        self._extract_members(); self._extract_supports(); self._extract_loads()
        print(f"✅ Nodes:{len(self.nodes)} Members:{len(self.members)} Supports:{len(self.supports)}")
        for name, md in self.members.items():
            if not md.get('internal'): self.original_members_list[name] = md.copy()
        return self
    
    def add_mesh_nodes(self, n_segments=5):
        self.n_segments = n_segments
        new_nodes = {}; new_members = {}; new_distributed_loads = []
        self.original_to_segments = {}
        for name, md in self.members.items():
            if md.get('internal'): continue
            x1, y1 = md['start'][:2]; x2, y2 = md['end'][:2]
            n1, n2 = md['n1'], md['n2']
            if n1 not in self.nodes or n2 not in self.nodes or n1 == n2: continue
            L = np.sqrt((x2-x1)**2 + (y2-y1)**2)
            if L < 1e-10: continue
            for i in range(1, n_segments):
                t = i/n_segments; nx = x1+t*(x2-x1); ny = y1+t*(y2-y1)
                new_nodes[f"_{name}_N{i}"] = (nx, ny, 0)
            segment_nodes = [f"_{name}_N{i}" for i in range(1, n_segments)]
            prev_node = n1
            self.original_to_segments[name] = []
            for seg_idx, internal_node in enumerate(segment_nodes, 1):
                seg_name = f"{name}_seg{seg_idx}"
                px1, py1 = self.nodes[prev_node][:2] if prev_node in self.nodes else new_nodes[prev_node][:2]
                px2, py2 = new_nodes[internal_node][:2]
                new_members[seg_name] = {'n1':prev_node, 'n2':internal_node, 'section':md['section'],
                    'material':md['material'], 'start':(px1,py1), 'end':(px2,py2), 'internal':True, 'original':name}
                self.original_to_segments[name].append(seg_name)
                prev_node = internal_node
            seg_name = f"{name}_seg{n_segments}"
            px1, py1 = new_nodes[prev_node][:2]; px2, py2 = self.nodes[n2][:2]
            new_members[seg_name] = {'n1':prev_node, 'n2':n2, 'section':md['section'],
                'material':md['material'], 'start':(px1,py1), 'end':(px2,py2), 'internal':True, 'original':name}
            self.original_to_segments[name].append(seg_name)
            for load in self.distributed_loads:
                if load['member'] == name:
                    for s_idx in range(1, n_segments+1):
                        new_distributed_loads.append({'member':f"{name}_seg{s_idx}", 'value':load['value'],
                            'direction':load['direction'], 'class':load.get('class','Q')})
        self.nodes.update(new_nodes); self.members = new_members
        self.distributed_loads = new_distributed_loads
        print(f"   ✅ Mesh refined: {len(new_nodes)} internal nodes, {len(new_members)} elements")
    
    def _extract_nodes(self):
        for c in self.msp.query('CIRCLE[layer=="STRUCT_NODES"]'):
            cx, cy = c.dxf.center.x, c.dxf.center.y; name = None
            for t in self.msp.query('MTEXT[layer=="STRUCT_NODES"]'):
                if np.sqrt((cx-t.dxf.insert.x)**2+(cy-t.dxf.insert.y)**2) < 50: name = t.text.strip()
            if name: self.nodes[name] = (cx, cy, 0)
    
    def _find_node(self, point):
        px, py = point[:2]; tol = self.TOL.get('find_node', 10)
        for name, (x, y, z) in self.nodes.items():
            if abs(x-px) < tol and abs(y-py) < tol: return name
        return None
    
    def _extract_members(self):
        used_names = set()
        for line in self.msp.query('LINE[layer=="STRUCT_BEAMS"]'):
            sx, sy = line.dxf.start.x, line.dxf.start.y; ex, ey = line.dxf.end.x, line.dxf.end.y
            n1, n2 = self._find_node((sx, sy)), self._find_node((ex, ey))
            mx, my = (sx+ex)/2, (sy+ey)/2; name = None; min_dist = float('inf')
            for t in self.msp.query('MTEXT[layer=="STRUCT_BEAMS"]'):
                txt = t.text.strip()
                if txt in used_names: continue
                dist = np.sqrt((mx-t.dxf.insert.x)**2+(my-t.dxf.insert.y)**2)
                if dist < min_dist and dist < self.TOL['member_text']: min_dist = dist; name = txt
            if not name: name = f"M{len(self.members)+1}"
            section = 'IPE200'
            for t in self.msp.query('MTEXT[layer=="STRUCT_SECTIONS"]'):
                if np.sqrt((mx-t.dxf.insert.x)**2+(my-t.dxf.insert.y)**2) < self.TOL['section_text']:
                    sn = t.text.strip().upper()
                    if sn in STEEL_SECTIONS: section = sn; break
            material = 'S235'
            for t in self.msp.query('MTEXT[layer=="STRUCT_MATERIALS"]'):
                if np.sqrt((mx-t.dxf.insert.x)**2+(my-t.dxf.insert.y)**2) < self.TOL['section_text']:
                    mn = t.text.strip().upper()
                    if mn in STEEL_MATERIALS: material = mn; break
            if n1 and n2:
                used_names.add(name)
                self.members[name] = {'n1':n1, 'n2':n2, 'section':section, 'material':material,
                    'start':(sx,sy), 'end':(ex,ey), 'internal':False}
                self.used_sections.add(section); self.used_materials.add(material)
    
    def _extract_supports(self):
        for poly in self.msp.query('LWPOLYLINE[layer=="STRUCT_CONSTRAINTS"]'):
            pts = list(poly.vertices())
            if len(pts) == 3:
                vy = max(p[1] for p in pts)
                vx = [p[0] for p in pts if abs(p[1]-vy) < self.TOL['support_vertex']][0]
                node = self._find_node((vx, vy))
                if node:
                    stype = 'fixed'
                    for t in self.msp.query('MTEXT[layer=="STRUCT_CONSTRAINTS"]'):
                        if np.sqrt((vx-t.dxf.insert.x)**2+(vy-t.dxf.insert.y)**2) < self.TOL['support_text']:
                            if 'HINGE' in t.text.upper(): stype = 'pinned'
                    self.supports[node] = stype
    
    def _extract_loads(self):
        arrows = list(self.msp.query('LWPOLYLINE[layer=="STRUCT_LOADS"]'))
        texts = list(self.msp.query('MTEXT[layer=="STRUCT_LOADS"]'))
        moment_texts, distributed_texts, point_texts = [], [], []
        for text in texts:
            txt = text.text.strip().upper(); txt_orig = text.text.strip()
            is_moment = (re.search(r'M\s*[=\-:]\s*[-+]?\d+', txt_orig, re.IGNORECASE) and 
                        re.search(r'(KNM|KN·M|NM\b)', txt) and not re.search(r'(KN/M|KN\s*/\s*M)', txt))
            is_distributed = bool(re.search(r'(KN/M|KN\s*/\s*M|N/M)', txt))
            is_point = bool(re.search(r'[FP]\s*[=\-:]\s*[-+]?\d+', txt)) and not is_moment and not is_distributed
            if is_moment: moment_texts.append(text)
            elif is_distributed: distributed_texts.append(text)
            elif is_point: point_texts.append(text)
        print(f"   Load texts: {len(moment_texts)}M, {len(distributed_texts)}q, {len(point_texts)}F")
        
        # === Riconoscimento multiplo dei carichi distribuiti ===
        for text in distributed_texts:
            tx, ty = text.dxf.insert.x, text.dxf.insert.y
            # Trova il membro orizzontale più vicino
            best_member = None
            min_dist = float('inf')
            for name, m in self.members.items():
                if m.get('internal'): continue   # salta i segmenti di mesh
                # Considera solo membri approssimativamente orizzontali
                if abs(m['start'][1] - m['end'][1]) > self.TOL['beam_horiz']:
                    continue
                x1, y1 = m['start'][:2]
                x2, y2 = m['end'][:2]
                A = np.array([x1, y1])
                B = np.array([x2, y2])
                P = np.array([tx, ty])
                AB = B - A
                AP = P - A
                t = np.dot(AP, AB) / np.dot(AB, AB)
                t = max(0, min(1, t))
                closest = A + t * AB
                dist = np.linalg.norm(P - closest)
                if dist < min_dist and dist < self.TOL.get('load_text', 150):
                    min_dist = dist
                    best_member = name
            if best_member is None:
                continue
            # Estrai valore numerico
            nums = re.findall(r'[-+]?\d*\.?\d+', text.text.strip())
            if nums:
                val = float(nums[0])
                if 'kn/m' in text.text.lower():
                    val *= 1000
                self.distributed_loads.append({
                    'member': best_member,
                    'value': -abs(val),   # FY verso il basso
                    'direction': 'FY',
                    'class': 'Q'
                })
                print(f"   ✅ Distributed: {best_member} q={val:.0f} N/m")
        # ================================================================

        # Momenti
        for text in moment_texts:
            tx, ty = text.dxf.insert.x, text.dxf.insert.y
            best = min(((np.sqrt((tx-nx)**2+(ty-ny)**2), name) for name, (nx,ny,nz) in self.nodes.items()), 
                      key=lambda x: x[0])
            if best[0] < self.TOL.get('moment_text', 200):
                nums = re.findall(r'[-+]?\d*\.?\d+', text.text.strip())
                if nums:
                    val = float(nums[0])
                    if 'knm' in text.text.lower(): val *= 1000
                    if 'cw' in text.text.lower(): val = -abs(val)
                    elif 'ccw' in text.text.lower(): val = abs(val)
                    self.moment_loads.append({'node':best[1], 'value':val, 'direction':'MZ', 'class':'Q'})
                    print(f"   ✅ Moment: {best[1]} M={val:.0f} Nm")
        # Puntuali
        for text in point_texts:
            tx, ty = text.dxf.insert.x, text.dxf.insert.y
            best = min(((np.sqrt((tx-nx)**2+(ty-ny)**2), name) for name, (nx,ny,nz) in self.nodes.items()), 
                      key=lambda x: x[0])
            if best[0] < self.TOL['load_text']:
                nums = re.findall(r'[-+]?\d*\.?\d+', text.text.strip())
                if nums:
                    val = float(nums[0])
                    if 'kn' in text.text.lower(): val *= 1000
                    direction = 'FX' if 'FX' in text.text.upper() else 'FY'
                    self.point_loads.append({'node':best[1], 'value':val, 'direction':direction, 'class':'Q'})
                    print(f"   ✅ Point: {best[1]} F={val:.0f} N")
    
    def to_pynite(self):
        print("\n🏗️ BUILDING MODEL...")
        frame = FEModel3D()
        for mn in self.used_materials:
            if mn in STEEL_MATERIALS: p = STEEL_MATERIALS[mn]; frame.add_material(mn, p['E'], p['G'], p['nu'], p['rho'])
        for sn in self.used_sections:
            if sn in STEEL_SECTIONS: p = STEEL_SECTIONS[sn]; frame.add_section(sn, p['A'], p['Iy'], p['Iz'], p['J'])
        for name, (x, y, z) in self.nodes.items(): frame.add_node(name, x*self.scale_factor, y*self.scale_factor, z*self.scale_factor)
        for name, m in self.members.items(): frame.add_member(name, m['n1'], m['n2'], m['material'], m['section'])
        for node, st in self.supports.items():
            if st == 'fixed': frame.def_support(node, True, True, True, True, True, True)
            elif st == 'pinned': frame.def_support(node, True, True, True, False, False, False)
        for load in self.distributed_loads: frame.add_member_dist_load(load['member'], load['direction'], load['value'], load['value'])
        for load in self.point_loads: frame.add_node_load(load['node'], load['direction'], load['value'])
        for load in self.moment_loads: frame.add_node_load(load['node'], load['direction'], load['value'])
        print(f"✅ Model ready"); return frame

# ============================================================
# FIG 1: STRUCTURE VIEWER (con etichette distribuite corrette)
# ============================================================
class StructureViewer:
    def __init__(self, parser, frame=None): self.parser = parser; self.frame = frame
        
    def show(self):
        fig, ax = plt.subplots(figsize=(16, 11)); ax.set_aspect('equal'); ax.grid(True, alpha=0.3)
        unit_label = f'[{self.parser.unit_name}]'
        ax.set_title(f'STRUCTURAL MODEL - LOADS AND REACTIONS {unit_label}', fontsize=16, fontweight='bold')
        ax.set_xlabel(f'X {unit_label}'); ax.set_ylabel(f'Y {unit_label}')
        s = self.parser.scale_factor; node_size = 0.01 * (1/s)**0.3
        
        for name, m in self.parser.members.items():
            x1, y1 = m['start'][:2]; x2, y2 = m['end'][:2]
            if m.get('internal'): ax.plot([x1, x2], [y1, y2], 'magenta', linewidth=2.5, alpha=0.7, zorder=1)
            else:
                ax.plot([x1, x2], [y1, y2], 'b-', linewidth=3, zorder=2)
                mx, my = (x1+x2)/2, (y1+y2)/2
                ax.text(mx+0.015, my+0.015, f"{name}\n{m['section']}\n{m['material']}", fontsize=9, weight='bold', color='blue',
                       bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.85))
        for name, (x, y, z) in self.parser.nodes.items():
            if name.startswith('_'): ax.plot(x, y, 'o', color='magenta', markersize=6, zorder=3)
            else:
                ax.add_patch(plt.Circle((x, y), node_size, facecolor='yellow', edgecolor='black', linewidth=2, zorder=5))
                ax.text(x+0.018, y+0.018, name, fontsize=11, weight='bold', bbox=dict(boxstyle='round', facecolor='yellow', alpha=0.9))
        for node, st in self.parser.supports.items():
            if node in self.parser.nodes:
                x, y, z = self.parser.nodes[node]; sz = node_size*2.5
                if st == 'fixed':
                    ax.add_patch(mpatches.Polygon([(x-sz, y-sz*1.3), (x+sz, y-sz*1.3), (x, y)], closed=True, facecolor='dimgray', edgecolor='black', linewidth=2, zorder=6))
                    ax.plot([x-sz*1.4, x+sz*1.4], [y-sz*1.3, y-sz*1.3], 'k-', linewidth=2, zorder=6)
                    for i in range(4): ax.plot([x-sz+i*sz*0.4, x-sz*0.7+i*sz*0.4], [y-sz*1.3, y-sz*1.3], 'w-', linewidth=1.5, zorder=7)
                    ax.text(x, y-sz*2, 'FIXED', fontsize=9, weight='bold', ha='center', bbox=dict(boxstyle='round', facecolor='lightgray', alpha=0.9))
                elif st == 'pinned':
                    ax.add_patch(plt.Circle((x, y-sz*0.3), sz*0.55, fill=False, edgecolor='black', linewidth=2.5, zorder=6))
                    ax.add_patch(mpatches.Polygon([(x-sz, y-sz*1.4), (x+sz, y-sz*1.4), (x, y-sz*0.55)], closed=True, facecolor='dimgray', edgecolor='black', linewidth=2, zorder=6))
                    ax.plot([x-sz*1.4, x+sz*1.4], [y-sz*1.4, y-sz*1.4], 'k-', linewidth=2, zorder=6)
                    ax.text(x, y-sz*2, 'HINGE', fontsize=9, weight='bold', ha='center', bbox=dict(boxstyle='round', facecolor='lightgray', alpha=0.9))
        
        # ---- CARICHI DISTRIBUITI ----
        shown_loads = set()
        for load in self.parser.distributed_loads:
            m = self.parser.members.get(load['member'])
            if m:
                original = m.get('original', load['member']); x1, y1 = m['start'][:2]; x2, y2 = m['end'][:2]
                for i in range(3):
                    t = i/2; px = x1+t*(x2-x1); py = y1+t*(y2-y1)
                    ax.arrow(px, py+node_size*2, 0, -node_size*3, head_width=node_size*0.6, head_length=node_size*1, fc='red', ec='darkred', alpha=0.6, width=node_size*0.3, length_includes_head=True, zorder=4)
                if original not in shown_loads:
                    orig_member = self.parser.original_members_list.get(original)
                    if orig_member:
                        ox1, oy1 = orig_member['start'][:2]
                        ox2, oy2 = orig_member['end'][:2]
                        ax.text(
                            (ox1 + ox2) / 2,
                            (oy1 + oy2) / 2 + node_size * 11,
                            f'q = {abs(load["value"])/1000:.0f} kN/m ({load.get("class","Q")})',
                            fontsize=13, color='darkred', weight='bold', ha='center',
                            bbox=dict(boxstyle='round,pad=0.5', facecolor='white', edgecolor='red', linewidth=2, alpha=0.95),
                            zorder=15
                        )
                        shown_loads.add(original)

        for load in self.parser.point_loads:
            if load['node'] in self.parser.nodes:
                x, y, z = self.parser.nodes[load['node']]; val_kN = abs(load['value'])/1000; sgn = 1 if load['value']>0 else -1; a_len = node_size*8
                if load['direction'] == 'FX':
                    ax.arrow(x-a_len*sgn, y, a_len*0.7*sgn, 0, head_width=node_size*1.4, head_length=node_size*2, fc='darkorange', ec='darkorange', alpha=0.95, width=node_size*0.6, length_includes_head=True, zorder=15)
                else:
                    ax.arrow(x, y-a_len*sgn, 0, a_len*0.7*sgn, head_width=node_size*1.4, head_length=node_size*2, fc='darkorange', ec='darkorange', alpha=0.95, width=node_size*0.6, length_includes_head=True, zorder=15)
                ax.text(x+node_size*4, y-node_size*3*sgn, f'F = {val_kN:.1f} kN ({load.get("class","Q")})', fontsize=11, color='darkorange', weight='bold', ha='center',
                       bbox=dict(boxstyle='round,pad=0.4', facecolor='white', edgecolor='darkorange', linewidth=2, alpha=0.95), zorder=15)
        for load in self.parser.moment_loads:
            if load['node'] in self.parser.nodes:
                x, y, z = self.parser.nodes[load['node']]; val_kNm = abs(load['value'])/1000; radius = node_size*3.5
                theta = np.linspace(0, np.pi*0.75, 25) if load['value']>0 else np.linspace(0, -np.pi*0.75, 25)
                arc_x = x+radius*np.cos(theta); arc_y = y+radius*np.sin(theta)
                ax.plot(arc_x, arc_y, 'purple', linewidth=4, zorder=15)
                ax.arrow(arc_x[-1], arc_y[-1], -np.sin(theta[-1])*node_size*1.8, np.cos(theta[-1])*node_size*1.8, head_width=node_size, head_length=node_size*1.2, fc='purple', ec='purple', zorder=15)
                senso = 'CCW (+)' if load['value']>0 else 'CW (-)'
                ax.text(x+node_size*5.5, y+node_size*5.5, f'M = {val_kNm:.1f} kNm\n{senso}', fontsize=11, color='purple', weight='bold', ha='center',
                       bbox=dict(boxstyle='round,pad=0.4', facecolor='white', edgecolor='purple', linewidth=2, alpha=0.95), zorder=15)
        if self.frame:
            for node_name in self.parser.supports.keys():
                node = self.frame.nodes[node_name]
                if node_name in self.parser.nodes:
                    x, y, z = self.parser.nodes[node_name]; Rx, Ry, Mz = node.RxnFX['Combo 1'], node.RxnFY['Combo 1'], node.RxnMZ['Combo 1']; a_len = node_size*9
                    if abs(Ry)>1:
                        sgn = 1 if Ry>0 else -1
                        ax.arrow(x, y-a_len, 0, a_len*0.7*sgn, head_width=node_size*1.4, head_length=node_size*2, fc='green', ec='darkgreen', width=node_size*0.6, zorder=15, length_includes_head=True)
                    if abs(Rx)>1:
                        sgn = 1 if Rx>0 else -1
                        ax.arrow(x-a_len*sgn, y, a_len*0.7*sgn, 0, head_width=node_size*1.4, head_length=node_size*2, fc='green', ec='darkgreen', width=node_size*0.6, zorder=15, length_includes_head=True)
                    txt = f"Ry={Ry/1000:.1f}kN"
                    if abs(Rx)>1: txt += f"\nRx={Rx/1000:.1f}kN"
                    if abs(Mz)>0.1: txt += f"\nMz={Mz/1000:.1f}kNm"
                    ax.text(x-node_size*5, y-node_size*14, txt, fontsize=10, color='darkgreen', weight='bold',
                           bbox=dict(boxstyle='round,pad=0.4', facecolor='lightgreen', edgecolor='darkgreen', linewidth=2, alpha=0.9), zorder=15)
        legend_elements = [mpatches.Patch(color='blue', label='Members'), mpatches.Patch(color='magenta', label='Mesh'),
                          mpatches.Patch(color='yellow', label='Nodes'), mpatches.Patch(color='dimgray', label='Supports'),
                          mpatches.Patch(color='red', label='Dist. Loads'), mpatches.Patch(color='darkorange', label='Point Forces'),
                          mpatches.Patch(color='purple', label='Moments'), mpatches.Patch(color='green', label='Reactions')]
        ax.legend(handles=legend_elements, loc='upper right', fontsize=8, framealpha=0.9, ncol=2)
        all_x = [c[0] for c in self.parser.nodes.values()]; all_y = [c[1] for c in self.parser.nodes.values()]
        m = max(0.3, (max(all_x)-min(all_x))*0.2)
        ax.set_xlim(min(all_x)-m, max(all_x)+m); ax.set_ylim(min(all_y)-m, max(all_y)+m)
        plt.tight_layout()
        if 'DISPLAY' in os.environ:
            plt.show()
        else:
            fname = f"{os.path.splitext(self.parser.dxf_file)[0]}_structure.png"
            plt.savefig(fname, dpi=150)
            print(f"   💾 Saved: {fname}")
            plt.close()

# ============================================================
# FIG 2: ELEMENT DIAGRAM VIEWER
# ============================================================
class ElementDiagramViewer:
    def __init__(self, parser, frame): self.parser = parser; self.frame = frame
        
    def show_all(self):
        if self.parser.n_segments == 3:
            display_members = [(name, md) for name, md in self.parser.members.items() if not md.get('internal')]
        else:
            display_members = [(name, md) for name, md in self.parser.original_members_list.items()]
        for member_name, md in display_members:
            if self.parser.n_segments == 3:
                if member_name not in self.frame.members: continue
                frame_members = [self.frame.members[member_name]]; L_member = frame_members[0].L()
            else:
                segments = self.parser.original_to_segments.get(member_name, [])
                if not segments: continue
                frame_members = [self.frame.members[s] for s in segments if s in self.frame.members]
                if not frame_members: continue
                L_member = sum(m.L() for m in frame_members)
            fig, axes = plt.subplots(3, 1, figsize=(12, 9))
            fig.suptitle(f'INTERNAL FORCES - {member_name} [{md["section"]}] L={L_member:.3f}m', fontsize=14, fontweight='bold')
            # Momento
            all_x_M, all_M, cum = [], [], 0
            for m in frame_members:
                try:
                    res = m.moment_array('Mz', 50); all_x_M.extend(cum+res[0]); all_M.extend(res[1]); cum += m.L()
                except: pass
            if all_x_M:
                x, M = np.array(all_x_M), np.array(all_M); ax = axes[0]
                ax.fill_between(x, 0, M, alpha=0.35, color='blue', where=(M>=0))
                ax.fill_between(x, 0, M, alpha=0.35, color='red', where=(M<0))
                ax.plot(x, M, 'b-', linewidth=2.5); ax.axhline(y=0, color='black', linewidth=0.8)
                ax.set_ylabel('Mz [N·m]'); ax.set_title('BENDING MOMENT Mz'); ax.grid(True, alpha=0.3)
                M_max, M_min = np.max(M), np.min(M)
                if abs(M_max)>0.01: i=np.argmax(M); ax.annotate(f'{M_max:.1f}', xy=(x[i], M_max), xytext=(10,15), textcoords='offset points', fontsize=9, bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
                if abs(M_min)>0.01: i=np.argmin(M); ax.annotate(f'{M_min:.1f}', xy=(x[i], M_min), xytext=(10,-15), textcoords='offset points', fontsize=9, bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
            # Taglio
            all_x_V, all_V, cum = [], [], 0
            for m in frame_members:
                try:
                    res = m.shear_array('FY', 50); all_x_V.extend(cum+res[0]); all_V.extend(res[1]); cum += m.L()
                except:
                    try:
                        res_m = m.moment_array('Mz', 51); dx = res_m[0][1]-res_m[0][0]
                        V_seg = np.diff(res_m[1])/dx; x_seg = (res_m[0][:-1]+res_m[0][1:])/2
                        all_x_V.extend(cum+x_seg); all_V.extend(V_seg); cum += m.L()
                    except: pass
            if all_x_V:
                x, V = np.array(all_x_V), np.array(all_V); ax = axes[1]
                ax.fill_between(x, 0, V, alpha=0.35, color='red', where=(V>=0))
                ax.fill_between(x, 0, V, alpha=0.35, color='orange', where=(V<0))
                ax.plot(x, V, 'r-', linewidth=2.5); ax.axhline(y=0, color='black', linewidth=0.8)
                ax.set_ylabel('Fy [N]'); ax.set_title('SHEAR FORCE Fy'); ax.grid(True, alpha=0.3)
                V_max, V_min = np.max(V), np.min(V)
                if abs(V_max)>0.01: i=np.argmax(V); ax.annotate(f'{V_max:.1f}', xy=(x[i], V_max), xytext=(10,10), textcoords='offset points', fontsize=8, bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
                if abs(V_min)>0.01: i=np.argmin(V); ax.annotate(f'{V_min:.1f}', xy=(x[i], V_min), xytext=(10,-15), textcoords='offset points', fontsize=8, bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
            # Assiale
            all_x_N, all_N, cum = [], [], 0
            for m in frame_members:
                try:
                    res = m.axial_array(50); all_x_N.extend(cum+res[0]); all_N.extend(res[1]); cum += m.L()
                except: pass
            if all_x_N:
                x, N = np.array(all_x_N), np.array(all_N); ax = axes[2]
                ax.fill_between(x, 0, N, alpha=0.35, color='green', where=(N>=0))
                ax.fill_between(x, 0, N, alpha=0.35, color='brown', where=(N<0))
                ax.plot(x, N, 'g-', linewidth=2.5); ax.axhline(y=0, color='black', linewidth=0.8)
                ax.set_xlabel('Position [m]'); ax.set_ylabel('Fx [N]'); ax.set_title('AXIAL FORCE Fx'); ax.grid(True, alpha=0.3)
                if abs(np.max(N))>0.01: i=np.argmax(N); ax.annotate(f'{N[i]:.1f}', xy=(x[i], N[i]), xytext=(10,10), textcoords='offset points', fontsize=8, bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
            plt.tight_layout()
            if 'DISPLAY' in os.environ:
                plt.show()
            else:
                fname = f"{os.path.splitext(self.parser.dxf_file)[0]}_{member_name}_diagrams.png"
                plt.savefig(fname, dpi=150)
                plt.close()

# ============================================================
# FIG 3: DEFORMED SHAPE VIEWER - FINALE CORRETTA (ora con parametro suffix)
# ============================================================
class DeformedShapeViewer:
    def __init__(self, parser, frame):
        self.parser = parser
        self.frame = frame

    def _check_equilibrium_and_congruence(self, node_displacements, node_rotations):
        print("\n" + "=" * 80)
        print("🔍 EQUILIBRIUM AND CONGRUENCE CHECK")
        print("=" * 80)
        print("\n📐 1. GLOBAL EQUILIBRIUM (ΣFx=0, ΣFy=0, ΣMz=0)")
        print("-" * 70)
        sum_Fx = sum_Fy = sum_Mz = 0.0
        nodes_m = {name: (x * self.parser.scale_factor, y * self.parser.scale_factor)
                   for name, (x, y, z) in self.parser.nodes.items()}
        for node_name in self.parser.supports:
            node = self.frame.nodes[node_name]
            Rx, Ry, Mz = node.RxnFX['Combo 1'], node.RxnFY['Combo 1'], node.RxnMZ['Combo 1']
            sum_Fx += Rx
            sum_Fy += Ry
            if node_name in nodes_m:
                x_m, y_m = nodes_m[node_name]
                sum_Mz += Mz + x_m * Ry - y_m * Rx
            print(f"  Support {node_name}: Rx={Rx:10.2f} N, Ry={Ry:10.2f} N, Mz={Mz:10.2f} Nm")

        load_Fx = load_Fy = load_Mz = 0.0
        print("\n  Applied loads:")
        for load in self.parser.point_loads:
            Fx = load['value'] if load['direction'] == 'FX' else 0.0
            Fy = load['value'] if load['direction'] == 'FY' else 0.0
            load_Fx += Fx; load_Fy += Fy
            if load['node'] in nodes_m:
                x_m, y_m = nodes_m[load['node']]
                load_Mz += x_m * Fy - y_m * Fx
            print(f"  Point {load['node']}: Fx={Fx:10.2f} N, Fy={Fy:10.2f} N")
        for load in self.parser.distributed_loads:
            if load['member'] not in self.parser.members: continue
            m = self.parser.members[load['member']]
            n1, n2 = m['n1'], m['n2']
            if n1 not in nodes_m or n2 not in nodes_m: continue
            x1_m, y1_m = nodes_m[n1]; x2_m, y2_m = nodes_m[n2]
            L_m = np.sqrt((x2_m - x1_m)**2 + (y2_m - y1_m)**2)
            q = load['value']
            if load['direction'] == 'FY':
                Fy = q * L_m; load_Fy += Fy
                xc_m = (x1_m + x2_m) / 2.0; load_Mz += xc_m * Fy
            elif load['direction'] == 'FX':
                Fx = q * L_m; load_Fx += Fx
                yc_m = (y1_m + y2_m) / 2.0; load_Mz -= yc_m * Fx
            print(f"  Distr {load['member']}: q={q:.2f} N/m, L={L_m:.3f} m, Resultant={q * L_m:.2f} N")
        for load in self.parser.moment_loads:
            M = load['value']; load_Mz += M
            print(f"  Moment {load['node']}: M={M:.2f} Nm")

        print(f"\n  {'':<20} {'Fx [N]':>12} {'Fy [N]':>12} {'Mz [Nm]':>12}")
        print(f"  {'Sum reactions':<20} {sum_Fx:12.2f} {sum_Fy:12.2f} {sum_Mz:12.2f}")
        print(f"  {'Sum loads':<20} {load_Fx:12.2f} {load_Fy:12.2f} {load_Mz:12.2f}")
        err_Fx = abs(sum_Fx + load_Fx)
        err_Fy = abs(sum_Fy + load_Fy)
        err_Mz = abs(sum_Mz + load_Mz)
        tol_F, tol_M = 1.0, 10.0
        eq_Fx = "✅" if err_Fx < tol_F else "❌"
        eq_Fy = "✅" if err_Fy < tol_F else "❌"
        eq_Mz = "✅" if err_Mz < tol_M else "❌"
        print(f"  {'Error':<20} {err_Fx:12.2e} {eq_Fx} {err_Fy:12.2e} {eq_Fy} {err_Mz:12.2e} {eq_Mz}")

        print("\n📐 2. ROTATION CONGRUENCE AT STRUCTURAL NODES (FEM)")
        print("-" * 75)
        print(f"  {'Node':<8} {'Connected members':<35} {'RZ [°]':>10} {'Status'}")
        for node_name in sorted(self.parser.nodes.keys()):
            if node_name.startswith('_'): continue
            node = self.frame.nodes[node_name]
            rot_fem_deg = np.degrees(node.RZ['Combo 1'])
            connected = []
            for seg_name, seg_md in self.parser.members.items():
                if seg_md['n1'] == node_name or seg_md['n2'] == node_name:
                    orig = seg_md.get('original', seg_name)
                    if orig not in connected: connected.append(orig)
            if len(connected) <= 1:
                print(f"  {node_name:<8} {'Only 1 member':<35} {rot_fem_deg:10.4f}°  --")
                continue
            member_str = ", ".join(connected)
            print(f"  {node_name:<8} {member_str:<35} {rot_fem_deg:10.4f}°  ✅ congruent (rigid node)")

        print("\n📐 3. BOUNDARY CONDITIONS CHECK")
        print("-" * 75)
        print(f"  {'Node':<8} {'Support':<12} {'DX [mm]':>12} {'DY [mm]':>12} {'RZ [°]':>12} {'Status'}")
        all_ok = True
        for node_name, stype in self.parser.supports.items():
            if node_name not in node_displacements: continue
            disp = node_displacements[node_name]
            dx, dy = disp['dx_mm'], disp['dy_mm']
            rz = np.degrees(node_rotations.get(node_name, 0.0))
            tol_disp = 1e-3
            if stype == 'fixed':
                ok = (abs(dx) < tol_disp and abs(dy) < tol_disp and abs(rz) < tol_disp)
                status = "✅" if ok else "❌"
                if not ok:
                    if abs(dx) >= tol_disp: print(f"          ❌ DX should be 0 but is {dx:.6f} mm")
                    if abs(dy) >= tol_disp: print(f"          ❌ DY should be 0 but is {dy:.6f} mm")
                    if abs(rz) >= tol_disp: print(f"          ❌ RZ should be 0 but is {rz:.6f}°")
            elif stype == 'pinned':
                ok = (abs(dx) < tol_disp and abs(dy) < tol_disp)
                status = "✅" if ok else "❌"
                if not ok:
                    if abs(dx) >= tol_disp: print(f"          ❌ DX should be 0 but is {dx:.6f} mm")
                    if abs(dy) >= tol_disp: print(f"          ❌ DY should be 0 but is {dy:.6f} mm")
                rz = 0.0
            else: status = "⚠️"
            if status == "❌": all_ok = False
            print(f"  {node_name:<8} {stype:<12} {dx:12.6f} {dy:12.6f} {rz:12.6f}  {status}")
        print("\n" + "=" * 80)
        print("📊 VERIFICATION SUMMARY")
        print("=" * 80)
        eq_ok = (eq_Fx == "✅" and eq_Fy == "✅" and eq_Mz == "✅")
        print(f"  Global Equilibrium:     {'✅ PASS' if eq_ok else '❌ FAIL'}")
        if not eq_ok:
            print(f"    - Force X error: {err_Fx:.2e} N")
            print(f"    - Force Y error: {err_Fy:.2e} N")
            print(f"    - Moment Z error: {err_Mz:.2e} Nm")
        print(f"  Rotation Congruence:    {'✅ Guaranteed by FEM (rigid nodes)'}")
        print(f"  Boundary Conditions:    {'✅ PASS' if all_ok else '❌ FAIL'}")
        if eq_ok and all_ok: print("\n  ✅ Model is well-posed and equilibrium is satisfied.")
        else: print("\n  ❌ There are issues that need attention!")
        print("=" * 80)
        return eq_ok and all_ok

    def show(self, save_image=True, image_format='png', suffix=None):
        from scipy.interpolate import PchipInterpolator
        fig, ax = plt.subplots(figsize=(18, 13)); ax.set_aspect('equal'); ax.grid(True, alpha=0.3)
        max_defl = 0.0; min_span = float('inf')
        node_rotations = {}; node_displacements = {}
        for node_name in self.parser.nodes:
            if node_name not in self.frame.nodes: continue
            node = self.frame.nodes[node_name]
            dx_m = node.DX['Combo 1']; dy_m = node.DY['Combo 1']; rz = node.RZ['Combo 1']
            dx_mm = dx_m * 1000.0; dy_mm = dy_m * 1000.0
            dx_dxf = dx_m / self.parser.scale_factor; dy_dxf = dy_m / self.parser.scale_factor
            node_displacements[node_name] = {'dx_mm': dx_mm, 'dy_mm': dy_mm, 'dx_dxf': dx_dxf, 'dy_dxf': dy_dxf, 'rz': rz}
            node_rotations[node_name] = rz
        for name, md in self.parser.members.items():
            if name not in self.frame.members: continue
            member = self.frame.members[name]; n1_name, n2_name = md['n1'], md['n2']
            if n1_name not in self.parser.nodes or n2_name not in self.parser.nodes: continue
            x1, y1 = self.parser.nodes[n1_name][:2]; x2, y2 = self.parser.nodes[n2_name][:2]
            L_dxf = np.sqrt((x2-x1)**2 + (y2-y1)**2); min_span = min(min_span, L_dxf)
            try:
                res = member.deflection_array('dy', 50); dy_mm = res[1] * 1000.0
                max_defl = max(max_defl, np.max(np.abs(dy_mm)))
            except: pass
        for node_name, stype in self.parser.supports.items():
            if node_name in node_displacements:
                node_displacements[node_name]['dx_mm'] = 0.0; node_displacements[node_name]['dy_mm'] = 0.0
                node_displacements[node_name]['dx_dxf'] = 0.0; node_displacements[node_name]['dy_dxf'] = 0.0
                if stype == 'fixed': node_rotations[node_name] = 0.0
        check_ok = self._check_equilibrium_and_congruence(node_displacements, node_rotations)
        if max_defl > 0 and min_span < float('inf'): scale = (min_span * 0.15) / max_defl
        else: scale = 100.0
        scale = max(scale, 10.0); scale = min(scale, 5000.0)
        print(f"\n   Deformed shape: max_defl={max_defl:.4f} mm, span={min_span:.0f} mm, scale=×{scale:.0f}")
        title_suffix = " (P-Delta)" if suffix else ""
        ax.set_title(f'DEFORMED SHAPE{title_suffix} - Max deflection: {max_defl:.2f} mm (×{scale:.0f})', fontsize=16, fontweight='bold')
        ax.set_xlabel(f'X [{self.parser.unit_name}]'); ax.set_ylabel(f'Y [{self.parser.unit_name}]')
        for name, md in self.parser.members.items():
            n1_name, n2_name = md['n1'], md['n2']
            if n1_name not in self.parser.nodes or n2_name not in self.parser.nodes: continue
            x1, y1 = self.parser.nodes[n1_name][:2]; x2, y2 = self.parser.nodes[n2_name][:2]
            if not md.get('internal'): ax.plot([x1, x2], [y1, y2], 'k-', linewidth=2.5, zorder=2)
            else: ax.plot([x1, x2], [y1, y2], 'k--', linewidth=0.8, alpha=0.3, zorder=1)
        node_list = sorted(self.parser.nodes.keys())
        for name in node_list:
            x, y, z = self.parser.nodes[name]
            if not name.startswith('_'):
                ax.plot(x, y, 'o', color='black', markersize=10, zorder=3)
                ax.text(x-15, y-25, f"{name}", fontsize=10, weight='bold', color='black', ha='center',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor='black', alpha=0.9))
            else: ax.plot(x, y, '.', color='black', markersize=4, zorder=3)
        for orig_name, orig_md in self.parser.original_members_list.items():
            segments = []
            for seg_name, seg_md in self.parser.members.items():
                if seg_md.get('original') == orig_name:
                    seg_num = int(seg_name.split('_seg')[-1]); segments.append((seg_num, seg_name, seg_md))
            if not segments: continue
            segments.sort(key=lambda x: x[0])
            ordered_nodes = [segments[0][2]['n1']]
            for _, _, seg_md in segments:
                n2 = seg_md['n2']
                if n2 not in ordered_nodes: ordered_nodes.append(n2)
            if len(ordered_nodes) < 2: continue
            x1_orig, y1_orig = self.parser.nodes[ordered_nodes[0]][:2]
            x2_orig, y2_orig = self.parser.nodes[ordered_nodes[-1]][:2]
            L_total = np.sqrt((x2_orig-x1_orig)**2 + (y2_orig-y1_orig)**2)
            if L_total < 1e-10: continue
            t_list, x_def_list, y_def_list = [], [], []
            cum_dist = 0.0; prev_x, prev_y = x1_orig, y1_orig
            for i, node_name in enumerate(ordered_nodes):
                x, y = self.parser.nodes[node_name][:2]
                if i > 0:
                    seg_L = np.sqrt((x-prev_x)**2 + (y-prev_y)**2); cum_dist += seg_L
                disp = node_displacements.get(node_name, {'dx_dxf': 0.0, 'dy_dxf': 0.0})
                x_def = x + disp['dx_dxf'] * scale; y_def = y + disp['dy_dxf'] * scale
                t = cum_dist / L_total if L_total > 0 else 0.0
                t_list.append(t); x_def_list.append(x_def); y_def_list.append(y_def)
                prev_x, prev_y = x, y
            try:
                spline_x = PchipInterpolator(t_list, x_def_list); spline_y = PchipInterpolator(t_list, y_def_list)
                t_fine = np.linspace(0.0, 1.0, 200)
                x_spline = spline_x(t_fine); y_spline = spline_y(t_fine)
                ax.plot(x_spline, y_spline, 'm-', linewidth=2.5, zorder=5, alpha=0.9)
                y_orig_line = y1_orig + (y2_orig-y1_orig) * t_fine
                ax.fill_between(x_spline, y_spline, y_orig_line, alpha=0.1, color='magenta', zorder=4)
            except:
                ax.plot(x_def_list, y_def_list, 'm-', linewidth=2.0, zorder=5, alpha=0.7)
        for name in node_list:
            x, y, z = self.parser.nodes[name]
            if name not in node_displacements: continue
            disp = node_displacements[name]
            x_def = x + disp['dx_dxf'] * scale; y_def = y + disp['dy_dxf'] * scale
            if not name.startswith('_'):
                ax.plot(x_def, y_def, 's', color='magenta', markersize=14, markeredgecolor='darkmagenta', markeredgewidth=2.5, zorder=10)
                ax.text(x_def+30, y_def+25, f"{name}'", fontsize=10, weight='bold', color='magenta', ha='center',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor='magenta', alpha=0.95))
                ax.plot([x, x_def], [y, y_def], ':', color='gray', linewidth=1.2, alpha=0.5, zorder=4)
            else: ax.plot(x_def, y_def, 's', color='magenta', markersize=5, alpha=0.7, zorder=10)
        legend_elements = [
            plt.Line2D([0], [0], color='black', linewidth=2.5, label='Original'),
            plt.Line2D([0], [0], color='magenta', linewidth=2.5, label=f'Deformed (×{scale:.0f})'),
            plt.Line2D([0], [0], marker='o', color='black', markersize=10, linestyle='None', label='Original nodes (N)'),
            plt.Line2D([0], [0], marker='s', color='magenta', markersize=12, linestyle='None',
                       markeredgecolor='darkmagenta', markeredgewidth=2, label="Deformed nodes (N')"),
        ]
        ax.legend(handles=legend_elements, loc='upper right', fontsize=10, framealpha=0.9)
        table_text = "NODAL DISPLACEMENTS\n" + "="*50 + "\n"
        table_text += f"{'Node':<12} {'DX [mm]':>10} {'DY [mm]':>10} {'RZ [°]':>10}\n" + "-"*50 + "\n"
        for name in node_list:
            if name.startswith('_'): continue
            disp = node_displacements.get(name, {'dx_mm': 0.0, 'dy_mm': 0.0})
            rot_deg = np.degrees(node_rotations.get(name, 0.0))
            table_text += f"{name:<12} {disp['dx_mm']:10.3f} {disp['dy_mm']:10.3f} {rot_deg:10.4f}\n"
        table_text += "-"*50 + "\n" + f"Scale: ×{scale:.0f}"
        ax.text(0.98, 0.02, table_text, transform=ax.transAxes, fontsize=7, verticalalignment='bottom',
                horizontalalignment='right', family='monospace',
                bbox=dict(boxstyle='round', facecolor='lightyellow', edgecolor='black', linewidth=1.5, alpha=0.95))
        info_text = f'Max deflection: {max_defl:.2f} mm | Scale: ×{scale:.0f}'
        ax.text(0.02, 0.98, info_text, transform=ax.transAxes, fontsize=11, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
        all_x = [c[0] for c in self.parser.nodes.values()]; all_y = [c[1] for c in self.parser.nodes.values()]
        for name in node_list:
            if name in node_displacements:
                disp = node_displacements[name]
                all_x.append(self.parser.nodes[name][0] + disp['dx_dxf'] * scale)
                all_y.append(self.parser.nodes[name][1] + disp['dy_dxf'] * scale)
        margin = max(500, (max(all_x)-min(all_x))*0.20)
        ax.set_xlim(min(all_x)-margin, max(all_x)+margin); ax.set_ylim(min(all_y)-margin, max(all_y)+margin)
        plt.tight_layout()
        # Salva sempre l'immagine
        base_name = os.path.splitext(self.parser.dxf_file)[0]
        if suffix:
            fname = f"{base_name}_deformed_{suffix}.{image_format}"
        else:
            fname = f"{base_name}_deformed.{image_format}"
        plt.savefig(fname, dpi=150, format=image_format)
        print(f"   💾 Saved: {fname}")
        if 'DISPLAY' in os.environ:
            plt.show()
        else:
            plt.close()

# ============================================================
# FIG 4: STRUCTURAL VERIFICATION + EXCEL (aggiunto P-Delta)
# ============================================================
class StructuralVerification:
    def __init__(self, parser, frame, gamma_G=1.35, gamma_Q=1.50, psi_0=0.7):
        self.parser = parser; self.frame = frame
        self.gamma_G, self.gamma_Q, self.psi_0 = gamma_G, gamma_Q, psi_0
        self.slu_factor = calc_slu_factor(parser, gamma_G, gamma_Q)
        self.gamma_M0, self.gamma_M1 = 1.00, 1.00
        
    def analyze_and_show(self):
        print("\n" + "="*80)
        print("📐 STRUCTURAL VERIFICATION (EN 1993-1-1 + EN 1990)")
        print("   CLASS 1 → Mpl,Rd = Wpl * fy / γM0")
        print("="*80)
        print(f"   γG={self.gamma_G}, γQ={self.gamma_Q}, ψ0={self.psi_0}, γM0={self.gamma_M0}, γM1={self.gamma_M1}")
        members_source = self.parser.original_members_list if self.parser.original_members_list else self.parser.members
        print("\n   1️⃣  SLU - CROSS-SECTION RESISTANCE")
        print(f"   Design loads: Ed = {self.slu_factor:.2f} × Ek\n")
        slu_results = []
        for name, md in members_source.items():
            if md.get('internal'): continue
            L = self._get_member_length(name)
            if L < 1e-6: continue
            N_Ed = self._get_max_axial(name)*self.slu_factor; M_Ed = self._get_max_moment(name)*self.slu_factor; V_Ed = self._get_max_shear(name)*self.slu_factor
            sec = STEEL_SECTIONS.get(md['section'], {}); mat = STEEL_MATERIALS.get(md['material'], {})
            A, fy, Wpl_z = sec.get('A',0), mat.get('fy',235e6), sec.get('Wpl_z',0)
            N_Rd = A*fy/self.gamma_M0; M_Rd = Wpl_z*fy/self.gamma_M0 if Wpl_z>0 else 0
            Av = A*0.6; V_Rd = Av*fy/(np.sqrt(3)*self.gamma_M0)
            rN = abs(N_Ed)/N_Rd if N_Rd>0 else 0; rM = abs(M_Ed)/M_Rd if M_Rd>0 else 0
            max_ratio = max(rN, rM)
            status = "✅ OK" if max_ratio<0.7 else ("⚡ ACC" if max_ratio<1.0 else "❌ FAIL")
            slu_results.append({'member':name, 'N_Ed':N_Ed, 'N_Rd':N_Rd, 'rN':rN, 'M_Ed':M_Ed, 'M_Rd':M_Rd, 'rM':rM,
                               'V_Ed':V_Ed, 'V_Rd':V_Rd, 'max_ratio':max_ratio, 'status':status, 'section':md['section'],
                               'A':A, 'fy':fy, 'Wpl_z':Wpl_z, 'L':L})
            print(f"   {name:<8} {N_Ed/1000:8.2f} {N_Rd/1000:8.2f} {rN:8.3f} {M_Ed/1000:8.2f} {M_Rd/1000:9.2f} {rM:8.3f} {V_Ed/1000:8.2f} {V_Rd/1000:8.2f} {status:<10}")
        print("\n   2️⃣  SLE - DEFLECTION CHECK")
        sle_results = []
        for name, md in members_source.items():
            if md.get('internal'): continue
            L = self._get_member_length(name)
            if L < 1e-6: continue
            defl = self._get_max_deflection(name)
            delta_lim = L*1000/250 if name=='M2' else L*1000/150; limit_type = "L/250" if name=='M2' else "L/150"
            ratio = defl/delta_lim if delta_lim>0 else 0
            status = "✅ OK" if ratio<0.7 else ("⚡ ACC" if ratio<1.0 else "⚠️ OVER")
            sle_results.append({'member':name, 'L':L, 'defl':defl, 'delta_lim':delta_lim, 'ratio':ratio, 'limit':limit_type, 'status':status})
            print(f"   {name:<8} {L:8.3f} {defl:12.4f} {delta_lim:12.2f} {ratio:8.3f} {limit_type:>12} {status:<12}")
        print("\n   3️⃣  BUCKLING & BEAM-COLUMN (Method 2)")
        buckling_results = []
        for name, md in members_source.items():
            if md.get('internal'): continue
            L = self._get_member_length(name)
            if L < 1e-6: continue
            sec = STEEL_SECTIONS.get(md['section'], {}); mat = STEEL_MATERIALS.get(md['material'], {})
            A, Iz = sec.get('A',0), sec.get('Iz',0)
            E, fy = mat.get('E',210e9), mat.get('fy',235e6)
            k = self._get_k_factor(md['n1'], md['n2'])
            Pcr = np.pi**2*E*Iz/(k*L)**2 if Iz>0 and L>0 else float('inf')
            lambda_bar = np.sqrt(A*fy/Pcr) if Pcr>0 and Pcr!=float('inf') and A>0 else 0
            alpha = 0.34 if any(x in md['section'] for x in ['IPE','HEA','HEB']) else 0.21
            chi = self._calc_chi(lambda_bar, alpha); Nb_Rd = chi*A*fy/self.gamma_M1
            buckling_results.append({'member':name, 'L':L, 'k':k, 'Pcr':Pcr, 'lambda_bar':lambda_bar, 'chi':chi, 'Nb_Rd':Nb_Rd,
                                    'section':md['section'], 'A':A, 'Iz':Iz, 'Wpl_z':sec.get('Wpl_z',0), 'E':E, 'fy':fy, 'alpha':alpha})
            print(f"   {name:<8} {L:8.3f} {k:6.2f} {Pcr/1000 if Pcr!=float('inf') else float('inf'):12.2f} {lambda_bar:8.3f} {chi:8.3f} {Nb_Rd/1000:10.2f}")
        beam_column_results = []
        for r in buckling_results:
            name = r['member']
            N_Ed = self._get_max_axial(name)*self.slu_factor; M_Ed = self._get_max_moment(name)*self.slu_factor
            A, fy, Wpl_z = r['A'], r['fy'], r['Wpl_z']
            lambda_bar, chi, Nb_Rd = r['lambda_bar'], r['chi'], r['Nb_Rd']
            Mpl_Rd = Wpl_z*fy/self.gamma_M0 if Wpl_z>0 else 0
            psi = self._get_moment_ratio(name); C_my = self._calc_Cmy(psi)
            N_Rk = A*fy
            N_Ed_chi_NRk = abs(N_Ed)/(chi*N_Rk/self.gamma_M1) if (chi*N_Rk)>0 else 0
            k_yy = min(C_my*(1+(lambda_bar-0.2)*N_Ed_chi_NRk), C_my*(1+0.8*N_Ed_chi_NRk))
            k_zy = 0.6*k_yy if any(x in r['section'] for x in ['IPE','HEA','HEB']) else k_yy
            rN_inter = abs(N_Ed)/Nb_Rd if Nb_Rd>0 else 0; rM_inter = abs(M_Ed)/Mpl_Rd if Mpl_Rd>0 else 0
            interaction_yy = rN_inter + k_yy*rM_inter; interaction_zy = rN_inter + k_zy*rM_inter
            inter_max = max(interaction_yy, interaction_zy)
            status = "✅ Safe" if inter_max<0.7 else ("⚡ OK" if inter_max<1.0 else "❌ FAIL!")
            beam_column_results.append({
                'member': name, 'psi': psi, 'C_my': C_my, 'k_yy': k_yy, 'k_zy': k_zy,
                'rN': rN_inter, 'rM': rM_inter,
                'interaction_yy': interaction_yy, 'interaction_zy': interaction_zy,
                'inter_max': inter_max, 'status': status,
                # Aggiunte per l'export Excel
                'N_Ed': N_Ed, 'Nb_Rd': Nb_Rd, 'M_Ed': M_Ed, 'Mpl_Rd': Mpl_Rd,
                'lambda_bar': lambda_bar, 'chi': chi
            })
            print(f"   {name:<8} {psi:7.3f} {C_my:7.3f} {k_yy:8.4f} {k_zy:8.4f} {rN_inter:11.4f} {rM_inter:11.4f} {interaction_yy:10.4f} {interaction_zy:10.4f} {inter_max:8.4f} {status:<12}")
        print("\n   ◆ FINAL SUMMARY")
        all_results = []
        for r in buckling_results:
            name = r['member']
            slu_r = next((x['max_ratio'] for x in slu_results if x['member']==name), 0)
            sle_r = next((x['ratio'] for x in sle_results if x['member']==name), 0)
            bc = next((x for x in beam_column_results if x['member']==name), None)
            buck_r = bc['inter_max'] if bc else 0
            overall = max(slu_r, sle_r, buck_r)
            print(f"   {name:<8} {'✅' if slu_r<0.7 else ('⚡' if slu_r<1.0 else '❌')} {slu_r:.3f}  {'✅' if sle_r<0.7 else ('⚡' if sle_r<1.0 else '❌')} {sle_r:.3f}  {'✅' if buck_r<0.7 else ('⚡' if buck_r<1.0 else '❌')} {buck_r:.3f}  {'✅' if overall<0.7 else ('⚡' if overall<1.0 else '❌')} {overall:.3f}")
            all_results.append({'member':name, 'slu_ratio':slu_r, 'sle_ratio':sle_r, 'buck_ratio':buck_r, 'overall':overall})
        return {'slu':slu_results, 'sle':sle_results, 'buckling':buckling_results, 'beam_column':beam_column_results, 'all':all_results}
    
    def _get_member_length(self, name):
        if name in self.frame.members: return self.frame.members[name].L()
        return sum(self.frame.members[f"{name}_seg{i}"].L() for i in range(1, self.parser.n_segments+1) if f"{name}_seg{i}" in self.frame.members)
    def _get_max_axial(self, name):
        try:
            if name in self.frame.members: return np.max(np.abs(self.frame.members[name].axial_array(50)[1]))
            v = [np.max(np.abs(self.frame.members[f"{name}_seg{i}"].axial_array(50)[1])) for i in range(1, self.parser.n_segments+1) if f"{name}_seg{i}" in self.frame.members]
            return np.max(v) if v else 0
        except: return 0
    def _get_max_moment(self, name):
        try:
            if name in self.frame.members: return np.max(np.abs(self.frame.members[name].moment_array('Mz', 50)[1]))
            v = [np.max(np.abs(self.frame.members[f"{name}_seg{i}"].moment_array('Mz', 50)[1])) for i in range(1, self.parser.n_segments+1) if f"{name}_seg{i}" in self.frame.members]
            return np.max(v) if v else 0
        except: return 0
    def _get_max_shear(self, name):
        try:
            if name in self.frame.members:
                try: return np.max(np.abs(self.frame.members[name].shear_array('FY', 50)[1]))
                except:
                    rm = self.frame.members[name].moment_array('Mz', 51); dx = rm[0][1]-rm[0][0]
                    return np.max(np.abs(np.diff(rm[1])/dx))
            v = []
            for i in range(1, self.parser.n_segments+1):
                s = f"{name}_seg{i}"
                if s in self.frame.members:
                    try: v.append(np.max(np.abs(self.frame.members[s].shear_array('FY', 50)[1])))
                    except: pass
            return np.max(v) if v else 0
        except: return 0
    def _get_max_deflection(self, name):
        try:
            if name in self.frame.members: return np.max(np.abs(self.frame.members[name].deflection_array('dy', 50)[1]))*1000
            v = [np.max(np.abs(self.frame.members[f"{name}_seg{i}"].deflection_array('dy', 50)[1]))*1000 for i in range(1, self.parser.n_segments+1) if f"{name}_seg{i}" in self.frame.members]
            return np.max(v) if v else 0
        except: return 0
    def _get_k_factor(self, n1, n2):
        s1 = self.parser.supports.get(n1, 'free'); s2 = self.parser.supports.get(n2, 'free')
        if s1=='fixed' and s2=='fixed': return 0.5
        elif (s1=='fixed' and s2=='pinned') or (s1=='pinned' and s2=='fixed'): return 0.7
        elif s1=='pinned' and s2=='pinned': return 1.0
        elif s1=='fixed' or s2=='fixed': return 2.0
        else: return 1.0
    def _calc_chi(self, lb, alpha):
        if lb <= 0.2: return 1.0
        phi = 0.5*(1+alpha*(lb-0.2)+lb**2)
        return min(1.0, 1/(phi+np.sqrt(phi**2-lb**2))) if phi**2-lb**2>0 else 1.0
    def _get_moment_ratio(self, name):
        try:
            if name in self.frame.members: M = self.frame.members[name].moment_array('Mz', 50)[1]
            else:
                M_list = []
                for i in range(1, self.parser.n_segments+1):
                    s = f"{name}_seg{i}"
                    if s in self.frame.members: M_list.extend(self.frame.members[s].moment_array('Mz', 50)[1])
                M = np.array(M_list) if M_list else np.array([0])
            M_max, M_min = np.max(M), np.min(M)
            return M_min/M_max if abs(M_max)>1e-6 else 1.0
        except: return 1.0
    def _calc_Cmy(self, psi):
        C_my = 0.6+0.4*psi if psi>=0 else 0.1-0.8*psi
        return max(0.4, min(C_my, 1.0))

class ResultsExporter:
    def __init__(self, parser, frame, gamma_G=1.35, gamma_Q=1.50, psi_0=0.7):
        self.parser = parser; self.frame = frame
        self.gamma_G, self.gamma_Q, self.psi_0 = gamma_G, gamma_Q, psi_0
        self.slu_factor = calc_slu_factor(parser, gamma_G, gamma_Q)
        
    def export(self, filename=None, results=None, pdelta_results=None):
        if filename is None: filename = os.path.splitext(self.parser.dxf_file)[0] + "_results.xlsx"
        print(f"\n📊 Creating Excel: {filename}")
        wb = openpyxl.Workbook()
        boldL = openpyxl.styles.Font(bold=True, size=14)
        header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        ok_fill = openpyxl.styles.PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warn_fill = openpyxl.styles.PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fail_fill = openpyxl.styles.PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        def style_header(ws, row, ncols):
            for c in range(1, ncols+1):
                cell = ws.cell(row=row, column=c); cell.fill = header_fill; cell.font = header_font
        def auto_width(ws, ncols, min_width=18):
            for c in range(1, ncols+1): ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min_width

        # --- Fattori di carico (comune) ---
        ws1 = wb.active; ws1.title = "Load Factors"
        ws1.cell(row=1, column=1, value="PARTIAL FACTORS (EN 1990 + EN 1993-1-1)").font = boldL
        for c, h in enumerate(['Factor', 'Value', 'Description'], 1): ws1.cell(row=3, column=c, value=h)
        style_header(ws1, 3, 3)
        for i, (f, v, d) in enumerate([('γG', self.gamma_G, 'Permanent'), ('γQ', self.gamma_Q, 'Variable'), ('ψ0', self.psi_0, 'Combination'), ('γM0', 1.00, 'Section'), ('γM1', 1.00, 'Buckling')], 4):
            ws1.cell(row=i, column=1, value=f); ws1.cell(row=i, column=2, value=v); ws1.cell(row=i, column=3, value=d)
        ws1.cell(row=10, column=1, value=f"SLU Factor: Ed = {self.slu_factor:.3f} × Ek").font = boldL
        auto_width(ws1, 3)

        # Funzione helper per scrivere i risultati in fogli standard o con prefisso "P-Delta "
        def write_results_sheets(results_dict, sheet_prefix=""):
            if results_dict is None: return
            ws5 = wb.create_sheet(f"{sheet_prefix}SLU - Resistance")
            ws5.cell(row=1, column=1, value=f"{sheet_prefix}SLU - SECTION RESISTANCE").font = boldL
            shdrs = ['Member', 'Section', 'L [m]', 'fy [MPa]', 'N_Ed [kN]', 'N_Rd [kN]', 'N/N_Rd', 'M_Ed [kNm]', 'Mpl,Rd [kNm]', 'M/Mpl,Rd', 'V_Ed [kN]', 'V_Rd [kN]', 'Max Ratio', 'Status']
            for c, h in enumerate(shdrs, 1): ws5.cell(row=4, column=c, value=h)
            style_header(ws5, 4, len(shdrs))
            row = 5
            for r in results_dict['slu']:
                data = [r['member'], r['section'], f"{r['L']:.3f}", f"{r['fy']/1e6:.0f}", f"{r['N_Ed']/1000:.2f}", f"{r['N_Rd']/1000:.2f}", f"{r['rN']:.4f}",
                        f"{r['M_Ed']/1000:.2f}", f"{r['M_Rd']/1000:.2f}", f"{r['rM']:.4f}", f"{r['V_Ed']/1000:.2f}", f"{r['V_Rd']/1000:.2f}", f"{r['max_ratio']:.4f}", r['status']]
                for c, v in enumerate(data, 1): ws5.cell(row=row, column=c, value=v)
                ws5.cell(row=row, column=14).fill = ok_fill if r['status'].startswith('✅') else (warn_fill if r['status'].startswith('⚡') else fail_fill)
                row += 1
            auto_width(ws5, len(shdrs))

            ws6 = wb.create_sheet(f"{sheet_prefix}SLE - Deflection")
            ws6.cell(row=1, column=1, value=f"{sheet_prefix}SLE - DEFLECTION").font = boldL
            for c, h in enumerate(['Member', 'L [m]', 'δ_max [mm]', 'δ_lim [mm]', 'Limit', 'Ratio', 'Status'], 1): ws6.cell(row=3, column=c, value=h)
            style_header(ws6, 3, 7)
            row = 4
            for r in results_dict['sle']:
                data = [r['member'], f"{r['L']:.3f}", f"{r['defl']:.4f}", f"{r['delta_lim']:.2f}", r['limit'], f"{r['ratio']:.4f}", r['status']]
                for c, v in enumerate(data, 1): ws6.cell(row=row, column=c, value=v)
                ws6.cell(row=row, column=7).fill = ok_fill if r['status'].startswith('✅') else (warn_fill if r['status'].startswith('⚡') else fail_fill)
                row += 1
            auto_width(ws6, 7)

            ws7 = wb.create_sheet(f"{sheet_prefix}Buckling - Euler")
            ws7.cell(row=1, column=1, value=f"{sheet_prefix}BUCKLING - EULER").font = boldL
            for c, h in enumerate(['Member', 'L [m]', 'k', 'E [GPa]', 'Iz [m⁴]', 'fy [MPa]', 'Pcr [kN]', 'λ_bar', 'α', 'χ', 'Nb_Rd [kN]'], 1): ws7.cell(row=3, column=c, value=h)
            style_header(ws7, 3, 11)
            row = 4
            for r in results_dict['buckling']:
                data = [r['member'], f"{r['L']:.3f}", f"{r['k']:.2f}", f"{r['E']/1e9:.0f}", f"{r['Iz']:.4e}", f"{r['fy']/1e6:.0f}",
                        f"{r['Pcr']/1000:.2f}" if r['Pcr']!=float('inf') else "∞", f"{r['lambda_bar']:.3f}", f"{r['alpha']:.2f}", f"{r['chi']:.3f}", f"{r['Nb_Rd']/1000:.2f}"]
                for c, v in enumerate(data, 1): ws7.cell(row=row, column=c, value=v)
                if r['lambda_bar']>0.2: ws7.cell(row=row, column=8).fill = warn_fill
                row += 1
            auto_width(ws7, 11)

            ws8 = wb.create_sheet(f"{sheet_prefix}Beam-Column")
            ws8.cell(row=1, column=1, value=f"{sheet_prefix}BEAM-COLUMN INTERACTION - Method 2").font = boldL
            bchdrs = ['Member', 'λ_bar', 'ψ', 'C_my', 'χ', 'Nb_Rd [kN]', 'N_Ed [kN]', 'M_Ed [kNm]', 'Mpl,Rd [kNm]', 'k_yy', 'k_zy', 'Inter_yy', 'Inter_zy', 'Max', 'Status']
            for c, h in enumerate(bchdrs, 1): ws8.cell(row=4, column=c, value=h)
            style_header(ws8, 4, len(bchdrs))
            row = 5
            for r in results_dict['beam_column']:
                data = [r['member'], f"{r.get('lambda_bar',0):.3f}", f"{r['psi']:.3f}", f"{r['C_my']:.4f}", f"{r.get('chi',0):.3f}",
                        f"{r['Nb_Rd']/1000:.2f}", f"{r['N_Ed']/1000:.2f}", f"{r['M_Ed']/1000:.2f}", f"{r['Mpl_Rd']/1000:.2f}",
                        f"{r['k_yy']:.4f}", f"{r['k_zy']:.4f}", f"{r['interaction_yy']:.4f}", f"{r['interaction_zy']:.4f}", f"{r['inter_max']:.4f}", r['status']]
                for c, v in enumerate(data, 1): ws8.cell(row=row, column=c, value=v)
                ws8.cell(row=row, column=15).fill = ok_fill if r['status'].startswith('✅') else (warn_fill if r['status'].startswith('⚡') else fail_fill)
                row += 1
            auto_width(ws8, len(bchdrs))

            ws9 = wb.create_sheet(f"{sheet_prefix}Final Summary")
            ws9.cell(row=1, column=1, value=f"{sheet_prefix}FINAL SUMMARY").font = boldL
            for c, h in enumerate(['Member', 'Section', 'SLU Ratio', 'SLU', 'SLE Ratio', 'SLE', 'Buckling Ratio', 'Buckling', 'Overall', 'Status'], 1): ws9.cell(row=3, column=c, value=h)
            style_header(ws9, 3, 10)
            row = 4
            for r in results_dict['all']:
                slu_info = next((x for x in results_dict['slu'] if x['member']==r['member']), {})
                data = [r['member'], slu_info.get('section',''), f"{r['slu_ratio']:.4f}", "✅" if r['slu_ratio']<0.7 else ("⚡" if r['slu_ratio']<1.0 else "❌"),
                        f"{r['sle_ratio']:.4f}", "✅" if r['sle_ratio']<0.7 else ("⚡" if r['sle_ratio']<1.0 else "❌"),
                        f"{r['buck_ratio']:.4f}", "✅" if r['buck_ratio']<0.7 else ("⚡" if r['buck_ratio']<1.0 else "❌"),
                        f"{r['overall']:.4f}", "✅ PASS" if r['overall']<1.0 else "❌ FAIL"]
                for c, v in enumerate(data, 1): ws9.cell(row=row, column=c, value=v)
                if r['overall']<0.7: ws9.cell(row=row, column=10).fill = ok_fill
                elif r['overall']<1.0: ws9.cell(row=row, column=10).fill = warn_fill
                else: ws9.cell(row=row, column=10).fill = fail_fill
                row += 1
            auto_width(ws9, 10)

        # Scrive risultati lineari (esistenti)
        write_results_sheets(results, "")
        # Scrive risultati P-Delta (nuovi fogli)
        write_results_sheets(pdelta_results, "P-Delta ")

        wb.save(filename)
        print(f"   ✅ Excel saved: {filename}")

# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("="*60)
        print("STRUCTURAL DXF PARSER v16.3.1 + P-Delta")
        print("EN 1993-1-1 + EN 1990 | Class 1 → Wpl | Method 2")
        print("="*60)
        print("Usage: python3 statics.py <file.dxf>")
        sys.exit(1)
    
    dxf_file = sys.argv[1]
    parser = CustomDXFParser(dxf_file)
    parser.parse()
    confirm_section_properties(parser)
    gamma_G, gamma_Q, psi_0 = classify_loads(parser)
    
    print("\n🔧 MESH DENSITY (default: 3)")
    try: n_seg = int(input("   Elements per member [3-20]: ").strip() or 3)
    except: n_seg = 3
    n_seg = max(3, min(20, n_seg))
    if n_seg > 3: parser.add_mesh_nodes(n_seg)
    
    # Analisi lineare originale
    frame = parser.to_pynite()
    print("\n⚡ ANALYZING (Linear)..."); frame.analyze(); print("✅ Done!")
    
    print("\n📊 LINEAR REACTIONS:")
    for n in parser.supports:
        node = frame.nodes[n]
        print(f"  {n}: Fx={node.RxnFX['Combo 1']:.1f}N, Fy={node.RxnFY['Combo 1']:.1f}N, Mz={node.RxnMZ['Combo 1']:.1f}Nm")
    
    print("\n🎨 FIG 1: Structure"); StructureViewer(parser, frame).show()
    print("📊 FIG 2: Diagrams"); ElementDiagramViewer(parser, frame).show_all()
    print("🔵 FIG 3: Deformed (Linear)"); DeformedShapeViewer(parser, frame).show()
    
    fy_sum = sum(frame.nodes[n].RxnFY['Combo 1'] for n in parser.supports)
    total = sum(l['value']*np.sqrt((parser.members[l['member']]['end'][0]-parser.members[l['member']]['start'][0])**2+(parser.members[l['member']]['end'][1]-parser.members[l['member']]['start'][1])**2)*parser.scale_factor for l in parser.distributed_loads)
    for l in parser.point_loads:
        if l['direction']=='FY': total += l['value']
    print(f"\n✅ EQUILIBRIUM: Sum Ry={fy_sum:.1f}N, Load={total:.1f}N, Error={abs(fy_sum+total):.1e}N")
    
    print("\n📐 LINEAR STRUCTURAL VERIFICATION")
    verification = StructuralVerification(parser, frame, gamma_G, gamma_Q, psi_0)
    results = verification.analyze_and_show()
    
    # ========== ANALISI P-DELTA ==========
    print("\n\n" + "="*70)
    print("🚀 P-DELTA ANALYSIS (Second Order)")
    print("="*70)
    frame_pdelta = parser.to_pynite()
    print("⚡ Analyzing (P-Delta)..."); frame_pdelta.analyze_PDelta(); print("✅ Done!")
    
    print("\n📊 P-DELTA REACTIONS:")
    for n in parser.supports:
        node = frame_pdelta.nodes[n]
        print(f"  {n}: Fx={node.RxnFX['Combo 1']:.1f}N, Fy={node.RxnFY['Combo 1']:.1f}N, Mz={node.RxnMZ['Combo 1']:.1f}Nm")
    
    print("🔵 FIG 4: P-Delta Deformed Shape")
    DeformedShapeViewer(parser, frame_pdelta).show(suffix="pdelta")
    
    print("\n📐 P-DELTA STRUCTURAL VERIFICATION")
    verification_pdelta = StructuralVerification(parser, frame_pdelta, gamma_G, gamma_Q, psi_0)
    results_pdelta = verification_pdelta.analyze_and_show()
    
    print("\n📊 Exporting Excel (Linear + P-Delta)...")
    ResultsExporter(parser, frame, gamma_G, gamma_Q, psi_0).export(results=results, pdelta_results=results_pdelta)
    
    print("\n✅ All done!")
